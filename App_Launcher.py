import sys
import docx
import time
import openpyxl
import threading
from tkinter import *
from tkinter import filedialog
from selenium import webdriver
from datetime import datetime
from cryptography.fernet import Fernet
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys 
from selenium.webdriver.support.ui import Select
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


#globla variables
listOfQuery = []
stopTheProgram = False
ii = 0

#loading the excel file
def getDataFromExcel(specificFilePath):
    wb = openpyxl.load_workbook(specificFilePath)
    sheet = wb.active
    data = []
    for i in range(2,sheet.max_row+1):
        temp = []
        temp.append(str(sheet.cell(row=i,column=10).value))
        temp.append(str(sheet.cell(row=i,column=14).value))
        temp.append(str(sheet.cell(row=i,column=16).value))
        temp.append(str(sheet.cell(row=i,column=17).value))
        temp.append(str(sheet.cell(row=i,column=18).value))
        temp.append(str(sheet.cell(row=i,column=15).value))
        temp.append(str(sheet.cell(row=i,column=22).value))
        temp.append(str(sheet.cell(row=i,column=23).value))
        temp.append(str(sheet.cell(row=i,column=7).value))
        data.append(temp)
    return data

#filling the output areas
######################################################
def fillOutputArea(outputBox,query):
    outputBox.insert(END,query)
######################################################


#file Loading from desktop
def loadExcelFile():
     global filePath
     latestTime.set("Wait Loading....")
     load_button.config(state=DISABLED,bg='red')
     filePath = filedialog.askopenfilename()
     global listOfQuery
     listOfQuery = getDataFromExcel(filePath)
     var = "Last Updated: " + str(time.strftime("%H:%M:%S", time.localtime()))
     latestTime.set(var)
     load_button.config(state=ACTIVE,bg=orig_color)

####################################################################

def getOnlyValue(specificFilePath):
    with open(specificFilePath,'r') as f_in:
        z = f_in.readline()
        return str(z).strip()
    
def getPerfectHoldingType(s):
    s  = s[0].lower() + s[1:]
    if(s[0] == 'i' and s[2] =='s'):
        s+='_1'
    return s
    
def getPhtCount(s,y):
    if(s == 'Individual'):
        if(s==y):
            return 1,1
        elif(s=='Joint'):
            return 0,1
        else:
            return 0,2
    elif(s == 'Joint'):
        if(s==y):
            return 1,1
        elif(s=='Individual'):
            return 1,0
        else:
            return 0,1 
    else :
        if(len(s)>10):
            return 1,1
        elif(s=='Individual'):
            return 2,0
        elif(s=='Joint'):
            return 1,0
    
def getPerfectGender(s):
    s  = s[0].lower() + s[1:]
    if(s[0] == 'i'):
        s+='_2'
    return s

def getPgCount(s,y):
    if(s == 'Male'):
        if(s==y):
            return 0,0
        elif(s=='Female'):
            return 0,1
        elif(s=='Transgender'):
            return 0,2
        else:
            return 0,0
    elif(s == 'Female'):
        if(s==y):
            return 1,1
        elif(s=='Male'):
            return 1,0
        elif(s=='Transgender'):
            return 0,1
        else:
            return 0,2
    elif(s == 'Transgender'):
        if(s==y):
            return 1,1
        elif(s=='Male'):
            return 2,0
        elif(s=='Female'):
            return 1,0
        else:
            return 0,1
    else:
        return 0,0


def getPerfectSoicalGroup(s):
    if(len(s)==2):
        s  = s[0].lower() + s[1].lower()
    else:
        s  = s[0].lower() + s[1:]
    if(s[0] == 'i'):
        s+='_3'
    return s

def getPsgCount(s,y):
    if(y == "Others"):
        if(s==y):
            return 1,1
        elif(s=="SC"):
            return 2,0
        elif(s=="ST"):
            return 1,0
    elif(y == "ST"):
        if(s==y):
            return 1,1
        elif(s=="SC"):
            return 1,0
        elif(s=="Others"):
            return 0,1
    elif(y == "SC"):
        if(s==y):
            return 0,0
        elif(s=="ST"):
            return 0,1
        elif(s=="Others"):
            return 0,2
    else:
        return 0,0    
    
def getleftright(a,b):
    decimal_value = float(str(a) + '.' + str(b))
    decimal_value = decimal_value/4
    decimal_value = str(decimal_value).split('.')
    return decimal_value[0],decimal_value[1]
#########################################################################################################################################
def getElement(driver,byWhat,strng,waitTime):
    wait = WebDriverWait(driver, waitTime)
    return wait.until(EC.presence_of_element_located((byWhat,strng)))

def operateElement(value,element,keysValue,waitTime):
    start_time = time.time()
    if(value == "click"):
        while time.time() - start_time < waitTime:
            try:
                element.click()
                break  
            except:
                time.sleep(0.5)
    elif(value == "send_keys"):
        while time.time() - start_time < waitTime:
            try:
                element.send_keys(keysValue)
                break
            except:
                time.sleep(0.5)

def decrypt_string(encrypted_string):
    key = 'ECPHuqGMo6QE2tcLElUX2GBmvOngpzFTbPAO09KMqdo='
    f = Fernet(key)
    decrypted = f.decrypt(encrypted_string)
    return decrypted.decode()
#########################################################################################################################################
def logic():
    global stopTheProgram
    try:
        username = username_entry.get()
        password = password_entry.get()

        #check if username valid 
        with open('./appData/allowed_users.txt','r') as f_in:
            text = f_in.read()
        text = decrypt_string(text)
        if(username not in text):
            labelAuthText.set("Unauthorized")
            with open('./memo/log.txt','a') as f:
                    f.write(f'unauthorized' + '\n')
            return
        else:
            labelAuthText.set("Authorized")

        service = Service('chromedriver.exe')
        driver = webdriver.Chrome(service=service) 
        driver.get("https://agcensus.gov.in/AgriCensus/")
        driver.maximize_window()
        ##############################################################################################
        element = driver.find_elements(By.CLASS_NAME,"hover-img")
        operateElement("click",element[1],"",10)
        driver.switch_to.window(driver.window_handles[-1]) 
        element = getElement(driver,By.ID,"state_list",10)
        select = Select(element)
        select.select_by_visible_text("15 Maharashtra")
        element = getElement(driver,By.ID,"user_id",10)
        operateElement("send_keys",element,username,10)
        element = getElement(driver,By.ID,"password",10)
        operateElement("send_keys",element,password,10)
        element = getElement(driver,By.ID,"captcha",10)
        text = element.text
        words = text.split()
        result = words[-1]

        if(stopTheProgram):
            driver.close()
            driver.quit()
            return
        
        element = getElement(driver,By.ID,"textbox",30)
        operateElement("send_keys",element,result,30)
        element = getElement(driver,By.ID,"Procced",30)
        operateElement("click",element,"",30)
        element = driver.find_elements(By.CSS_SELECTOR,"span.d-lg-flex.d-sm-inline-block.ms-lg-0.ms-3")
        operateElement("click",element[1],"",30)
        element = driver.find_elements(By.CSS_SELECTOR,"a.dropdown-item.fw-bold[onclick='L1_Entry()")
        operateElement("click",element[0],"",30)

        if(stopTheProgram):
            driver.close()
            driver.quit()
            return
        
        with open("./memo/log.txt", "w") as f:
            f.write("")

        global ii
        while(ii<len(listOfQuery)):
            try:
                if(stopTheProgram):
                    driver.close()
                    driver.quit()
                    return
                serialNumber = listOfQuery[ii][0]
                holdingType = listOfQuery[ii][1]
                gender = listOfQuery[ii][2]
                socialGroup = listOfQuery[ii][3]
                placeOfResidence = listOfQuery[ii][4]
                jointNumber = listOfQuery[ii][5]
                nameOfInstitution = listOfQuery[ii][6]
                villageName = listOfQuery[ii][7]
                sector = listOfQuery[ii][7]
                element = getElement(driver,By.ID,"sr_no",10)
                element.send_keys(Keys.BACKSPACE*4)
                operateElement("send_keys",element,serialNumber,10)
                element = getElement(driver,By.ID,"khata_no",10)
                operateElement("send_keys",element,"",10)

                element = getElement(driver,By.ID,"vlg_list",10)
                select = Select(element)
                select.select_by_visible_text(villageName)

                while(1==1):
                    time.sleep(1)
                    element = getElement(driver,By.CLASS_NAME,"loader",10)
                    if(element.is_displayed() == False):
                        break
                time.sleep(1)

                z = getPerfectHoldingType(holdingType)
                element = getElement(driver,By.ID,z,10)
                operateElement("send_keys",element,'\n',10)
                y = element.get_attribute("value")
                a,b = getPhtCount(holdingType,y)
                operateElement("send_keys",element,(Keys.ARROW_LEFT*a) + (Keys.ARROW_RIGHT*b),10)
                if(z == 'joint'):
                    element = getElement(driver,By.ID,"joint_no",10)
                    operateElement("send_keys",element,Keys.BACKSPACE*2,10)
                    operateElement("send_keys",element,jointNumber,10)

                time.sleep(1)
                z = getPerfectGender(gender)
                element = getElement(driver,By.ID,'male',10)
                operateElement("send_keys",element,'\n',10)
                y = element.get_attribute("value")
                a,b = getPgCount(gender,y)
                if(a==0 and b==0):
                    operateElement("send_keys",element,(Keys.ARROW_RIGHT*a)+(Keys.ARROW_LEFT*b),10)
                else :
                    operateElement("send_keys",element,(Keys.ARROW_LEFT)+(Keys.ARROW_RIGHT),10)


                z = getPerfectSoicalGroup(socialGroup)
                element = getElement(driver,By.ID,z,10)
                operateElement("send_keys",element,'\n',10)
                y = element.get_attribute("value")
                a,b = getPsgCount(socialGroup,y)
                if(a==0 and b==0):
                    operateElement("send_keys",element,(Keys.ARROW_RIGHT*a)+(Keys.ARROW_LEFT*b),10)
                else :  
                    operateElement("send_keys",element,(Keys.ARROW_LEFT)+(Keys.ARROW_RIGHT),10)
                    
                time.sleep(1)
                
                element = getElement(driver,By.ID,"place_resident",10)
                select = Select(element)
                select.select_by_value(placeOfResidence)

            
                if(nameOfInstitution != None and nameOfInstitution!='None' and nameOfInstitution!='' and nameOfInstitution!=' '):
                    element = getElement(driver,By.ID,"inst_name",10)
                    operateElement("send_keys",element,Keys.BACKSPACE*30,10)
                    operateElement("send_keys",element,nameOfInstitution,10)
                    
                    time.sleep(1)

                    element = getElement(driver,By.ID,"inst_sector",10)
                    select = Select(element)
                    select.select_by_value(sector)
                time.sleep(1)
                
                element = getElement(driver,By.ID,"tot_survey",10)
                z = element.get_attribute("value")
                z = int(z)
                for j in range(0,z):
                    element1 = "area" + str(j+1)
                    element2 = "area" + str(j+1) + "_1"
                    element3 = "land_use_list"+str(j+1)
                    element4 = "operational_holder_list"+str(j+1)
                    a,b = 0,0
                    elements1 = getElement(driver,By.XPATH,f'//input[@name="{element1}"]',10)
                    elements2 = getElement(driver,By.XPATH,f'//input[@name="{element2}"]',10)
                    a = elements1.get_attribute("value")
                    b = elements2.get_attribute("value")
                    a,b = getleftright(a,b)
                    elements1.send_keys(Keys.BACKSPACE*4)
                    elements2.send_keys(Keys.BACKSPACE*4)
                    operateElement("send_keys",elements1,a,10)
                    operateElement("send_keys",elements2,b,10)
                    element = getElement(driver,By.XPATH,f'//select[@name="{element3}"]',10)
                    select = Select(element)
                    select.select_by_value("1")
                    element = getElement(driver,By.XPATH,f'//select[@name="{element4}"]',10)
                    select = Select(element)
                    select.select_by_value("1")
                driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                element = getElement(driver,By.XPATH,'//textarea[@name="remarks"]',10)
                operateElement("send_keys",element,"" + Keys.TAB + Keys.ENTER,10)
                time.sleep(1)
                countTime = 0
                while(1==1):
                    countTime+=1
                    time.sleep(1)
                    element = getElement(driver,By.CLASS_NAME,"loader",10)
                    if(element.is_displayed() == False):
                        break
                    if(countTime >= 50):
                        with open('./memo/log.txt','a') as f:
                            f.write(f'{serialNumber} server was busy retrying for others {datetime.now()}')
                        fillOutputArea(outputArea1,serialNumber)                      
                        driver.close()
                        driver.quit() 
                        return
                with open('./memo/log.txt','a') as f:
                    f.write(f'{serialNumber} done at {datetime.now()}')
                fillOutputArea(outputArea2,serialNumber)
                fillOutputArea(outputArea3,serialNumber)
            except Exception as e:
                with open('./memo/log.txt','a') as f:
                    f.write(f'failed for {serialNumber} with error {e} at {datetime.now()}')
                fillOutputArea(outputArea1,serialNumber)
                driver.close()
                driver.quit()
                return
            i+=1
            time.sleep(1)
    except Exception as e:
        with open('./memo/log.txt','a') as f:
            f.write(f'failed  with error {e} at {datetime.now()}')
        fillOutputArea(outputArea1,serialNumber)
        driver.close()
        driver.quit()
        return


nameOfThread = 'MyThread'

def startProgram():
    global stopTheProgram
    stopTheProgram = False
    startButton.config(state=DISABLED,bg='LIGHT GREEN')
    threading.Thread(target=logic).start()
    global ii
    if(ii<len(listOfQuery)):
         threading.Thread(target=logic,name = 'MyThread').start()
    stopButtonText.set("stop")
    stopButton.config(state=NORMAL,bg=orig_color)
    startButton.config(state=NORMAL,bg=orig_color)



def stopProgram():
    stopButtonText.set("stoping")
    stopButton.config(state=DISABLED,bg='red')
    global stopTheProgram
    global nameOfThread
    for thread in threading.enumerate():
        if(thread.name!='MainThread'):
            stopTheProgram = True
            thread.join()

    stopButtonText.set("stop")
    startButton.config(state=NORMAL,bg=orig_color)
    


root = Tk()
root.title("Auto_Survey")
root.iconbitmap('./appData/icon.ico')
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()
root.minsize(800,600)
root.geometry("800x600")
root.maxsize(800,600)
root.resizable(width=False, height=False)  



username_label = Label(root, text="Username:")
username_label.grid(row=0, column=0, padx=5, pady=5, sticky=W)

# Create entry field for username
username_entry = Entry(root)
username_entry.grid(row=0, column=1, padx=5, pady=5)

# Create label for password
password_label = Label(root, text="Password:")
password_label.grid(row=1, column=0, padx=5, pady=5, sticky=W)

# Create entry field for password
password_entry = Entry(root, show="*")
password_entry.grid(row=1, column=1, padx=5, pady=5)

#load button
load_button = Button(root, text="LOAD EXCEL FILE", command=loadExcelFile)
load_button.grid(row=2,column=1,padx=0,pady=0)
orig_color = load_button.cget("background")

latestTime = StringVar()
latestTime.set("Please Load The File")
modifiedLabel = Label(root,textvariable=latestTime).grid(row=2,column=0,padx=0)

#output labels
output_label1 = Label(root, text="FAIL")
output_label1.grid(row=4, column=1, padx=5, pady=0, sticky=W)

output_label2 = Label(root, text="SUCCESS")
output_label2.grid(row=4, column=9, padx=5, pady=0, sticky=W)

output_label3 = Label(root, text="MEMORY")
output_label3.grid(row=4, column=16, padx=5, pady=0, sticky=W)


#outputs
outputArea1 = Text(root, width=20, height=28)
outputArea1.grid(row=5, column=0, columnspan=4, padx=50, pady=0)

outputArea2 = Text(root, width=20, height=28)
outputArea2.grid(row=5, column=7, columnspan=4, padx=50, pady=0)

outputArea3 = Text(root, width=20, height=28)
outputArea3.grid(row=5, column=14, columnspan=4, padx=50, pady=0)

def on_modified1(event):
    textOfOutputArea = outputArea1.get('1.0', 'end')
    with open ("./memo/failure.txt","w") as f:
        f.write(textOfOutputArea)

def on_modified2(event):
    textOfOutputArea = outputArea2.get("1.0", "end")
    with open ("./memo/success.txt","w") as f:
        f.write(textOfOutputArea)

def on_modified3(event):
    textOfOutputArea = outputArea3.get("1.0", "end")
    with open ("./memo/memory.txt","w") as f:
        f.write(textOfOutputArea)

def on_change1(event):
    textOfOutputArea = username_entry.get()
    with open ("./memo/username.txt","w") as f:
        f.write(textOfOutputArea)

def on_change2(event):
    textOfOutputArea = password_entry.get()
    with open ("./memo/password.txt","w") as f:
        f.write(textOfOutputArea)

outputArea1.bind('<KeyRelease>', on_modified1)
outputArea2.bind('<KeyRelease>', on_modified2)
outputArea3.bind('<KeyRelease>', on_modified3)
username_entry.bind('<KeyRelease>', on_change1)
password_entry.bind('<KeyRelease>', on_change2)


with open ("./memo/failure.txt","r") as f:
    for line in f:
        fillOutputArea(outputArea1,line)

with open ("./memo/success.txt","r") as f:
    for line in f:
        fillOutputArea(outputArea2,line)

with open ("./memo/memory.txt","r") as f:
    for line in f:
        fillOutputArea(outputArea3,line)

with open ("./memo/username.txt","r") as f:
    for line in f:
        username_entry.insert(0,line)

with open ("./memo/password.txt","r") as f:
    for line in f:
        password_entry.insert(0,line)

#logButton
###########################################################################################
def log_button_click():
    # Create a new window
    log_window = Toplevel(root)
    log_window.title("Log Window")

    # Create a Text widget in the new window
    log_text = Text(log_window)
    log_text.pack()

    # Redirect the standard output to the Text widget
    with open("./memo/log.txt", "r") as f:
        for line in f:
            log_text.insert(END, line)
            log_text.insert(END,"\n")
    
###########################################################################################
###########################################################################################

startButton = Button(root, text="START", command=startProgram)
startButton.grid(row=0,column=16,padx=3,pady=5)


stopButtonText = StringVar()
stopButtonText.set("STOP")
stopButton = Button(root, textvariable=stopButtonText, command=stopProgram)
stopButton.grid(row=0,column=17,padx=3,pady=5)

labelAuthText = StringVar()
labelAuthText.set("")
labelAuth = Label(root,textvariable=labelAuthText)
labelAuth.grid(row=0, column=10, padx=0, pady=0, sticky=W)

logButton = Button(root, text="LOG", command=log_button_click)
logButton.grid(row=0,column=15,padx=3,pady=5)
root.mainloop()